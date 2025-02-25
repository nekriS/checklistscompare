# -*- coding: utf-8 -*-
import openpyxl
import os
import datetime

class options:
    def __init__(self, output_file_path = "default_name.xlsx", sch_allow = True, db_allow = True, pcb_allow = False, find_allow = False, checker_flow = False, open_file = False):
        self.output_file_path = output_file_path    # имя выходного файла
        self.sch_allow = sch_allow        # разрешение на проверку листов схемотехники
        self.db_allow = db_allow        # разрешение на проверку листов базы данных
        self.pcb_allow = pcb_allow        # разрешение на проверку листов pcb
        self.find_allow = find_allow        # разрешение на поиск позиций в чек листе
        self.checker_flow = checker_flow        # разрешение на перенос исполнителей
        self.open_file = open_file


def log(text, log_object):
    # Получаем сегодняшнюю дату в формате YYYY-MM-DD
    today_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # Создаем путь к каталогу и файлу
    log_directory = 'log'
    log_file_name = f'log_{today_date}.txt'
    log_file_path = os.path.join(log_directory, log_file_name)

    # Проверяем наличие каталога "log" и создаем его, если его нет
    if not os.path.exists(log_directory):
        os.makedirs(log_directory)

    # Формируем строку для записи: "дата-время > текст"
    current_time = datetime.datetime.now().strftime('%H:%M:%S')
    log_entry = f"{today_date} {current_time} > {text}"

    # Проверяем existence файла и записываем данные
    with open(log_file_path, 'a', encoding='utf-8') as file:
        file.write(log_entry+"\n")

    log_object.append(log_entry)


    print(log_entry)


def read_cell_range(sheet, start_row, end_row, start_col, end_col):
    """
    Считывает данные из указанного диапазона ячеек на листе Excel.

    :param sheet: Объект листа (openpyxl worksheet)
    :param start_row: Начальная строка (целое число)
    :param end_row: Конечная строка (целое число)
    :param start_col: Начальный столбец (целое число, A=1, B=2, ...)
    :param end_col: Конечный столбец (целое число)
    :return: Двумерный массив (список списков) с данными из ячеек
    """
    data = []
    for row in range(start_row, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)
    return data

def mas_no_mas(massive):
    out_massive = []
    for el in massive:
        if el[0] != None:
            out_massive.append(el[0])
    return out_massive


def findrow(sheet, number, start_row, end_row, column = 3):
    for row in range(start_row, end_row):
        if sheet.cell(row=row, column=column).value == number:
            return row
    return -1

def howmanyrows(sheet, start = 1):

    class rows_class:
        def __init__(self, start_row = -1, end_row = -1):
            self.start_row = start_row    # начальная строка
            self.end_row = end_row    # конечная строка

    rows = rows_class()
    row = start

    while (rows.end_row == -1) or (row == 10^3):
        cell_value = sheet.cell(row=row, column=1).value
        if (("Да" in str(cell_value))  and ("Нет" in str(cell_value)) and ("Не проверено" in str(cell_value))) or (str(cell_value) in "ДаНетНе проверено"):
            if rows.start_row == -1:
                rows.start_row = row
        elif rows.start_row != -1:
            rows.end_row = row - 1
        row += 1


    return rows

def compare(path_1, path_2, path_3, checker, options, log_object):
    # параметры в options: output_file_path, sch_allow, db_allow, pcb_allow, find_allow
    no = 0
    yes = 0
    warn = 0

    try:
        workbook = openpyxl.load_workbook(path_1)
        workbook_middle = openpyxl.load_workbook(path_2)
        workbook_end = openpyxl.load_workbook(path_3)
    except:
        log("Файлы не доступны для открытия!", log_object)
        return [no, yes, warn]

    sheet_names = workbook.sheetnames

    main_sheet = workbook.worksheets[0]

    rw = howmanyrows(main_sheet)
    #print(rw.start_row)
    #print(rw.end_row)

    main_sheet_end = workbook_end.worksheets[0]

    PNs = mas_no_mas(read_cell_range(main_sheet, rw.start_row, rw.end_row, 2, 2))
    CHs = mas_no_mas(read_cell_range(main_sheet, rw.start_row, rw.end_row, 6, 6))

    # перенос проверяющих
    if options.checker_flow:
        for row in range(rw.start_row, rw.end_row):
            cell_value = main_sheet.cell(row=row, column=1).value
            if cell_value != None:
                row_end = findrow(main_sheet_end, main_sheet.cell(row=row, column=2).value, 8, 60, 2)
                if row_end != -1:
                    main_sheet_end.cell(row=row_end, column=6).value = main_sheet.cell(row=row, column=6).value

    # определяем компоненты которые нужно проверить по параметру "кто проверяет"
    if checker != "all":
        for ch_n in reversed(range(len(CHs))):
            if CHs[ch_n] != checker:
                PNs.pop(ch_n)


    for sheet_name in sheet_names[1:]:

        sheet = workbook[sheet_name]  # Получение объекта листа по имени
        sheet_middle = workbook_middle[sheet_name]


        #print(rw.start_row)
        #print(rw.end_row)

        try:
            sheet_end = workbook_end[sheet_name]  # Получение объекта листа по имени
        except:
            log("Листа " + sheet_name + " не обнаружено!", log_object)
            warn += 1
            continue

        part_number = sheet["B1"].value

        if (" DB" in sheet_name) and options.db_allow:
            rw = howmanyrows(sheet)
            if part_number in PNs:

                #data = read_cell_range(sheet, 8, 54, 1, 4)
                sheet_end.cell(row=7, column=5).value = "1. Значение"
                sheet_end.cell(row=7, column=6).value = "2. Комментарий"

                for row in range(rw.start_row, rw.end_row):

                    cell_value = sheet.cell(row=row, column=1).value

                    if cell_value != None:

                        if cell_value == "Да":
                            value_first = sheet.cell(row=row, column=4).value
                            value_end = sheet_end.cell(row=row, column=4).value
                            if value_first == value_end:
                                sheet_end.cell(row=row, column=1).value = "Да"
                            else:
                                sheet_end.cell(row=row, column=1).value = "Нет"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row, column=6).value = sheet_middle.cell(row=row, column=2).value
                        elif cell_value == "Нет":
                            value_first = sheet.cell(row=row, column=4).value
                            value_end = sheet_end.cell(row=row, column=4).value
                            if value_first == value_end:
                                sheet_end.cell(row=row, column=1).value = "Нет"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row, column=6).value = sheet_middle.cell(row=row, column=2).value
                            else:
                                #sheet_end.cell(row=row, column=1).value = "Не проверено"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row, column=6).value = sheet_middle.cell(row=row, column=2).value


        if (" Sch" in sheet_name) and options.sch_allow:
            rw = howmanyrows(sheet, 14)
            if part_number in PNs:

                # проверка шапки
                sheet_end.cell(row=7, column=5).value = "1. Значение"
                sheet_end.cell(row=7, column=6).value = "1. Комментарий"
                for row in range(8, 12):

                    cell_value = sheet.cell(row=row, column=1).value

                    if cell_value != None:
                        if cell_value == "Да":
                            value_first = sheet.cell(row=row, column=4).value
                            value_end = sheet_end.cell(row=row, column=4).value
                            if value_first == value_end:
                                sheet_end.cell(row=row, column=1).value = "Да"
                            else:
                                sheet_end.cell(row=row, column=1).value = "Нет"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value

                                if "Проверяется соответствие " in sheet.cell(row=row, column=2).value:
                                    sheet_end.cell(row=row, column=6).value = sheet.cell(row=row, column=5).value
                                else:
                                    sheet_end.cell(row=row, column=6).value = str(sheet.cell(row=row, column=2).value) + " " + str(sheet.cell(row=row, column=5).value)

                        elif cell_value == "Нет":
                            value_first = sheet.cell(row=row, column=4).value
                            value_end = sheet_end.cell(row=row, column=4).value
                            if value_first == value_end:
                                sheet_end.cell(row=row, column=1).value = "Нет"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value

                                if "Проверяется соответствие " in sheet.cell(row=row, column=2).value:
                                    sheet_end.cell(row=row, column=6).value = sheet.cell(row=row, column=5).value
                                else:
                                    sheet_end.cell(row=row, column=6).value = str(sheet.cell(row=row, column=2).value) + " " + str(sheet.cell(row=row, column=5).value)

                            else:
                                sheet_end.cell(row=row, column=1).value = "Не проверено"
                                sheet_end.cell(row=row, column=5).value = sheet.cell(row=row, column=4).value

                                if "Проверяется соответствие " in sheet.cell(row=row, column=2).value:
                                    sheet_end.cell(row=row, column=6).value = sheet.cell(row=row, column=5).value
                                else:
                                    sheet_end.cell(row=row, column=6).value = str(sheet.cell(row=row, column=2).value) + " " + str(sheet.cell(row=row, column=5).value)



                # проверка распиновки, максимум 500 строк
                max_strings = 500
                sheet_end.cell(row=16, column=7).value = "1. Name"
                sheet_end.cell(row=16, column=8).value = "1. Type"
                sheet_end.cell(row=16, column=9).value = "2. Комментарий"

                for row in range(17, max_strings):

                    cell_value = sheet.cell(row=row, column=1).value

                    if cell_value != None:

                        if options.find_allow:
                            row_3 = findrow(sheet_end, sheet.cell(row=row, column=3).value, 17, max_strings)
                            if row_3 == -1:
                                log(str(part_number) + " > Параметр " + str(sheet.cell(row=row, column=3).value) + " не найден!", log_object)
                                warn += 1
                                continue
                        else:
                            row_3 = row
                        #print(row);
                        #print(row_3);
                        if cell_value == "Да":
                            value_name_first = sheet.cell(row=row, column=4).value
                            value_name_end = sheet_end.cell(row=row_3, column=4).value

                            value_type_first = sheet.cell(row=row, column=5).value
                            value_type_end = sheet_end.cell(row=row_3, column=5).value

                            if (value_name_first == value_name_end) and (value_type_first == value_type_end):
                                sheet_end.cell(row=row_3, column=1).value = "Да"
                                yes += 1
                            else:
                                sheet_end.cell(row=row_3, column=1).value = "Нет"
                                sheet_end.cell(row=row_3, column=7).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row_3, column=8).value = sheet_middle.cell(row=row, column=5).value
                                sheet_end.cell(row=row_3, column=9).value = sheet_middle.cell(row=row, column=2).value
                                no += 1
                        elif cell_value == "Нет":
                            value_name_first = sheet.cell(row=row, column=4).value
                            value_name_end = sheet_end.cell(row=row_3, column=4).value

                            value_type_first = sheet.cell(row=row, column=5).value
                            value_type_end = sheet_end.cell(row=row_3, column=5).value

                            if (value_name_first == value_name_end) and (value_type_first == value_type_end):
                                sheet_end.cell(row=row_3, column=1).value = "Нет"
                                sheet_end.cell(row=row_3, column=7).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row_3, column=8).value = sheet_middle.cell(row=row, column=5).value
                                sheet_end.cell(row=row_3, column=9).value = sheet_middle.cell(row=row, column=2).value
                                no += 1
                            else:
                                #sheet_end.cell(row=row, column=1).value = "Не проверено"
                                sheet_end.cell(row=row_3, column=7).value = sheet.cell(row=row, column=4).value
                                sheet_end.cell(row=row_3, column=8).value = sheet_middle.cell(row=row, column=5).value
                                sheet_end.cell(row=row_3, column=9).value = sheet_middle.cell(row=row, column=2).value

    #output_file_path = 'example_updated.xlsx'

    workbook_end.save(options.output_file_path)

    try:
        if options.open_file:
            os.startfile(options.output_file_path)
    except:
        log("Не удалось открыть полученный файл!", log_object)

    return [no, yes, warn]

if __name__ == "__main__":

    path_1 = ""
    path_2 = ""
    path_3 = ""

    checker = ""

    default_options = options()

    compare(path_1, path_2, path_3, checker, default_options)
